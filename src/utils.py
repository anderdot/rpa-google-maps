import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from config import (
    LATITUDE,
    LONGITUDE,
    RAIO,
    DATA_DIR,
    EXPORTS_DIR,
    TIMESTAMP
)

def normalizar_dados(dados, tipo_localizado):
    dados_normalizados = []
    for item in dados:
        nome_estabelecimento = item.get("name", "N/A")
        tipo = tipo_localizado
        nota_estabelecimento = item.get("rating", "N/A")
        quantidade_avaliacoes = item.get("user_ratings_total", "N/A")
        endereco_completo = item.get("vicinity", "N/A")

        dados_normalizados.append({
            "nome_estabelecimento": nome_estabelecimento,
            "tipo": tipo,
            "nota_estabelecimento": nota_estabelecimento,
            "quantidade_avaliacoes": quantidade_avaliacoes,
            "endereco_completo": endereco_completo
        })
    return dados_normalizados

def criar_nome_arquivo(extensao):
    lat = str(LATITUDE).replace(".", "-")
    lon = str(LONGITUDE).replace(".", "-")
    rai = str(RAIO)
    nome_arquivo = f"estabelecimentos_lat{lat}_lon{lon}_rai_{rai}_{TIMESTAMP}.{extensao}"

    return nome_arquivo

def gerar_json(dados):
    nome_arquivo = criar_nome_arquivo("json")
    with open(DATA_DIR / nome_arquivo, "w", encoding="utf-8") as arquivo:
        json.dump(dados, arquivo, indent=2, ensure_ascii=False)

def carregar_json(caminho):
    with open(caminho, "r", encoding="utf-8") as arquivo:
        return json.load(arquivo)

def gerar_excel(dados):
    nome_arquivo = criar_nome_arquivo("xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "estabelecimentos"

    if not dados:
        workbook.save(EXPORTS_DIR / nome_arquivo)
        return

    colunas = list(dados[0].keys())
    sheet.append(colunas)

    for item in dados:
        sheet.append([item.get(coluna) for coluna in colunas])

    for indice, coluna in enumerate(colunas, start=1):
        letra = get_column_letter(indice)
        tamanho_max = max(
            len(str(sheet[f"{letra}{linha}"].value or ""))
            for linha in range(1, sheet.max_row + 1)
        )
        sheet.column_dimensions[letra].width = tamanho_max + 2

    workbook.save(EXPORTS_DIR / nome_arquivo)

def exportar_dados(dados):
    gerar_json(dados)
    dados_json = carregar_json(DATA_DIR / criar_nome_arquivo("json"))
    gerar_excel(dados_json)