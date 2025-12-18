import json
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configuração do logger para este módulo
log = logging.getLogger(__name__)

def normalizar_dados(dados, tipo_localizado):
    """
    Normaliza os dados coletados da API do Google Maps.

    Args:
        dados (list): Lista de dicionários com os dados brutos.
        tipo_localizado (str): Tipo legível do estabelecimento localizado.

    Returns:
        list: Lista de dicionários com os dados normalizados.
    """
    dados_normalizados = []
    for item in dados:
        try:
            nome_estabelecimento = item.get("name", "N/A")
            tipo = tipo_localizado
            nota_estabelecimento = item.get("rating", "N/A")
            quantidade_avaliacoes = item.get("user_ratings_total", "N/A")
            endereco_completo = item.get("vicinity", "N/A")

            dados_normalizados.append({
                "nome_estabelecimento": nome_estabelecimento,
                "tipo": tipo,
                "nota_estabelecimento": nota_estabelecimento,
                "quantidade_avaliacoes": quantidade_avaliacoes,
                "endereco_completo": endereco_completo
            })
        except KeyError as erro:
            log.warning(f"Chave ausente ao normalizar dados: {erro}: {item}.")
    return dados_normalizados


def gerar_json(dados, caminho_json):
    """
    Gera um arquivo JSON a partir dos dados fornecidos.

    Args:
        dados (list): Lista de dicionários com os dados a serem exportados.
        caminho_json (str): Caminho do arquivo JSON a ser criado.
    """
    try:
        with open(caminho_json, "w", encoding="utf-8") as arquivo:
            json.dump(dados, arquivo, indent=2, ensure_ascii=False)
        log.debug(f"Json exportado com {len(dados)} registros: {caminho_json}")
    except OSError as erro:
        log.error(f"Erro ao gerar JSON: {erro}")
        raise


def carregar_json(caminho_json):
    """
    Carrega dados de um arquivo JSON.

    Args:
        caminho_json (str): Caminho do arquivo JSON a ser carregado.

    Returns:
        list: Lista de dicionários com os dados carregados do arquivo JSON.
    """
    try:
        with open(caminho_json, "r", encoding="utf-8") as arquivo:
            return json.load(arquivo)
    except FileNotFoundError:
        log.error(f"Arquivo JSON não encontrado: {caminho_json}")
        raise

    except json.JSONDecodeError as erro:
        log.error(f"JSON inválido: {erro}")
        raise


def gerar_excel(dados, caminho_excel):
    """
    Gera um arquivo Excel a partir dos dados fornecidos.

    Args:
        dados (list): Lista de dicionários com os dados a serem exportados.
        caminho_excel (str): Caminho do arquivo Excel a ser criado.
    """
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "estabelecimentos"

        if not dados:
            workbook.save(caminho_excel)
            return

        colunas = list(dados[0].keys())
        sheet.append(colunas)

        for item in dados:
            sheet.append([item.get(coluna) for coluna in colunas])

        # Ajusta automaticamente a largura das colunas
        for indice, coluna in enumerate(colunas, start=1):
            letra = get_column_letter(indice)

            # Calcula o maior tamanho de texto encontrado na coluna
            tamanho_max = max(
                len(str(sheet[f"{letra}{linha}"].value or ""))
                for linha in range(1, sheet.max_row + 1)
            )
            # Define a largura da coluna com uma margem extra para legibilidade
            sheet.column_dimensions[letra].width = tamanho_max + 2

        workbook.save(caminho_excel)
        log.debug(f"Excel exportado com {len(dados)} registros: {caminho_excel}")
    except Exception as erro:
        log.error(f"Erro ao gerar Excel: {erro}")
        raise


def exportar_dados(dados, caminhos_arquivos):
    """
    Exporta os dados para arquivos JSON e Excel.

    Args:
        dados (list): Lista de dicionários com os dados a serem exportados.
        caminhos_arquivos (dict): Dicionário contendo os caminhos dos arquivos JSON e Excel.
    """
    gerar_json(dados, caminhos_arquivos["json"])
    dados_json = carregar_json(caminhos_arquivos["json"])
    gerar_excel(dados_json, caminhos_arquivos["excel"])
